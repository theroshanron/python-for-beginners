import openpyxl as xl
from openpyxl import chart

from openpyxl.chart import BarChart, Reference

def  process(filename):
    wb=  xl.load_workbook(filename)
    sheet = wb['Sheet1']


# cell = sheet['a1']
# cell = sheet.cell(3,2)
# # sheet.cell(1,1)


# print(cell.value)
# print(sheet.max_row)
    for row in range(3, sheet.max_row + 1):
        cell = sheet.cell(row,3)
        corrected_price = (cell.value * 0.9)
        print(corrected_price)
        corrected_price_cell = sheet.cell(row,4)
        corrected_price_cell.value = corrected_price
        

    values = Reference(sheet, min_row = 2, max_row = sheet.max_row, min_col =4, max_col =4)
    print(values)

    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, 'e2')
    wb.save(filename)

process("Python_test.xlsx")



